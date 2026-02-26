from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


EVENT_COLORS = {
    "START": "#16a34a",
    "LID_OPEN": "#ef4444",
    "LID_CLOSE": "#3b82f6",
    "END": "#f59e0b",
}


def _fmt_dt(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value is None:
        return "--"
    return str(value)


def _fmt_num(value: Any, digits: int = 4) -> str:
    if value is None:
        return "--"
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)
    if digits <= 0:
        return f"{numeric:.0f}"
    return f"{numeric:.{digits}f}"


def _safe_text(value: Any) -> str:
    if value is None:
        return "--"
    text = str(value).strip()
    return text if text else "--"


def _build_summary_rows(batch_id: int, meta: dict, metrics: dict) -> list[tuple[str, str]]:
    return [
        ("Batch ID", str(batch_id)),
        ("Inicio", _fmt_dt(metrics.get("start_ts") or meta.get("start_ts"))),
        ("Fim", _fmt_dt(metrics.get("end_ts") or meta.get("end_ts"))),
        ("OP", _safe_text(meta.get("op"))),
        ("Operador", _safe_text(meta.get("operador"))),
        ("Observacoes", _safe_text(meta.get("observacoes"))),
        ("Duracao (s)", _fmt_num(metrics.get("duration_s"), 3)),
        ("Temp min/max/avg", f"{_fmt_num(metrics.get('min_temp'))} / {_fmt_num(metrics.get('max_temp'))} / {_fmt_num(metrics.get('avg_temp'))}"),
        (
            "Corrente min/max/avg",
            f"{_fmt_num(metrics.get('min_corrente'))} / {_fmt_num(metrics.get('max_corrente'))} / {_fmt_num(metrics.get('avg_corrente'))}",
        ),
        ("AUC corrente", _fmt_num(metrics.get("current_auc"), 6)),
        ("AUC temperatura", _fmt_num(metrics.get("temp_auc"), 6)),
        (
            "Energy proxy",
            f"{_fmt_num(metrics.get('energy_proxy'), 6)} (integral trapezoidal da corrente)",
        ),
    ]


def build_excel_report(
    batch_id: int,
    meta: dict,
    readings: list[dict],
    events: list[dict],
    metrics: dict,
    include_raw: bool = False,
) -> bytes:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Resumo"

    summary_ws.append(["Campo", "Valor"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)

    for label, value in _build_summary_rows(batch_id=batch_id, meta=meta, metrics=metrics):
        summary_ws.append([label, value])

    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 66

    events_ws = workbook.create_sheet("Eventos")
    events_ws.append(["timestamp", "tipo", "valor", "motivo"])
    for cell in events_ws[1]:
        cell.font = Font(bold=True)

    for event in events:
        events_ws.append(
            [
                _fmt_dt(event.get("ts")),
                _safe_text(event.get("type")),
                "" if event.get("value") is None else str(event.get("value")),
                _safe_text(event.get("reason")),
            ]
        )

    events_ws.column_dimensions["A"].width = 22
    events_ws.column_dimensions["B"].width = 14
    events_ws.column_dimensions["C"].width = 12
    events_ws.column_dimensions["D"].width = 30

    if include_raw:
        readings_ws = workbook.create_sheet("Leituras")
        readings_ws.append(["timestamp", "temperatura", "corrente", "start_signal", "lid_signal", "temp_raw", "corr_raw"])
        for cell in readings_ws[1]:
            cell.font = Font(bold=True)

        for row in readings:
            readings_ws.append(
                [
                    _fmt_dt(row.get("ts")),
                    row.get("temp"),
                    row.get("corrente"),
                    row.get("start_signal"),
                    row.get("lid_signal"),
                    row.get("temp_raw"),
                    row.get("corr_raw"),
                ]
            )

        readings_ws.column_dimensions["A"].width = 22
        readings_ws.column_dimensions["B"].width = 14
        readings_ws.column_dimensions["C"].width = 14
        readings_ws.column_dimensions["D"].width = 12
        readings_ws.column_dimensions["E"].width = 10
        readings_ws.column_dimensions["F"].width = 10
        readings_ws.column_dimensions["G"].width = 10

        if readings_ws.max_row >= 3:
            chart = LineChart()
            chart.title = "Temperatura e Corrente"
            chart.style = 2
            chart.height = 8
            chart.width = 18
            chart.y_axis.title = "Valor"
            chart.x_axis.title = "Timestamp"

            chart_data = Reference(readings_ws, min_col=2, max_col=3, min_row=1, max_row=readings_ws.max_row)
            categories = Reference(readings_ws, min_col=1, min_row=2, max_row=readings_ws.max_row)
            chart.add_data(chart_data, titles_from_data=True)
            chart.set_categories(categories)

            readings_ws.add_chart(chart, "I2")

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _build_plot_png(readings: list[dict], events: list[dict]) -> io.BytesIO:
    fig, ax_temp = plt.subplots(figsize=(10.8, 4.4), dpi=120)
    ax_corr = ax_temp.twinx()

    ts = [row.get("ts") for row in readings]
    temp = [row.get("temp") for row in readings]
    corrente = [row.get("corrente") for row in readings]

    if ts:
        ax_temp.plot(ts, temp, color="#ef4444", linewidth=1.6, label="Temperatura (C)")
        ax_corr.plot(ts, corrente, color="#3b82f6", linewidth=1.6, label="Corrente (A)")
        ax_temp.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
        fig.autofmt_xdate(rotation=30)
    else:
        ax_temp.text(0.5, 0.5, "Sem leituras para este batch", ha="center", va="center", transform=ax_temp.transAxes)

    ax_temp.set_ylabel("Temperatura (C)")
    ax_corr.set_ylabel("Corrente (A)")
    ax_temp.grid(alpha=0.25)

    for event in events:
        event_ts = event.get("ts")
        if not isinstance(event_ts, datetime):
            continue
        event_type = str(event.get("type") or "")
        color = EVENT_COLORS.get(event_type, "#6b7280")
        ax_temp.axvline(event_ts, color=color, linestyle="--", linewidth=1.0, alpha=0.85)

    handles_left, labels_left = ax_temp.get_legend_handles_labels()
    handles_right, labels_right = ax_corr.get_legend_handles_labels()
    if handles_left or handles_right:
        ax_temp.legend(handles_left + handles_right, labels_left + labels_right, loc="upper left", fontsize=8)

    fig.tight_layout()

    out = io.BytesIO()
    fig.savefig(out, format="png")
    plt.close(fig)
    out.seek(0)
    return out


def build_pdf_report(
    batch_id: int,
    meta: dict,
    readings: list[dict],
    events: list[dict],
    metrics: dict,
) -> bytes:
    out = io.BytesIO()

    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    normal = styles["BodyText"]

    story = []
    story.append(Paragraph(f"Relatorio automatico da batelada {batch_id}", styles["Title"]))
    story.append(Spacer(1, 0.3 * cm))

    header_table_data = [
        ["Batch ID", str(batch_id), "OP", _safe_text(meta.get("op"))],
        ["Inicio", _fmt_dt(metrics.get("start_ts") or meta.get("start_ts")), "Operador", _safe_text(meta.get("operador"))],
        ["Fim", _fmt_dt(metrics.get("end_ts") or meta.get("end_ts")), "Observacoes", _safe_text(meta.get("observacoes"))],
    ]
    header_table = Table(header_table_data, colWidths=[2.2 * cm, 5.2 * cm, 2.2 * cm, 8.2 * cm])
    header_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9ca3af")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 0.35 * cm))

    metrics_table_data = [
        ["Metrica", "Valor"],
        ["Duracao (s)", _fmt_num(metrics.get("duration_s"), 3)],
        ["Temp min/max/avg", f"{_fmt_num(metrics.get('min_temp'))} / {_fmt_num(metrics.get('max_temp'))} / {_fmt_num(metrics.get('avg_temp'))}"],
        [
            "Corrente min/max/avg",
            f"{_fmt_num(metrics.get('min_corrente'))} / {_fmt_num(metrics.get('max_corrente'))} / {_fmt_num(metrics.get('avg_corrente'))}",
        ],
        ["AUC corrente", _fmt_num(metrics.get("current_auc"), 6)],
        ["AUC temperatura", _fmt_num(metrics.get("temp_auc"), 6)],
        ["Energy proxy", f"{_fmt_num(metrics.get('energy_proxy'), 6)} (integral da corrente)"] ,
    ]

    metrics_table = Table(metrics_table_data, colWidths=[6.4 * cm, 11.2 * cm])
    metrics_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9ca3af")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.append(metrics_table)
    story.append(Spacer(1, 0.4 * cm))

    plot_buffer = _build_plot_png(readings=readings, events=events)
    plot_image = Image(plot_buffer, width=17.5 * cm, height=7.0 * cm)
    story.append(plot_image)
    story.append(Spacer(1, 0.3 * cm))

    events_table_data = [["timestamp", "tipo", "valor", "motivo"]]
    for event in events:
        events_table_data.append(
            [
                _fmt_dt(event.get("ts")),
                _safe_text(event.get("type")),
                "--" if event.get("value") is None else str(event.get("value")),
                _safe_text(event.get("reason")),
            ]
        )

    events_table = Table(events_table_data, colWidths=[4.1 * cm, 2.5 * cm, 2.0 * cm, 9.3 * cm])
    events_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fef3c7")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(events_table)
    story.append(Spacer(1, 0.2 * cm))

    story.append(
        Paragraph(
            "Observacao: energy_proxy corresponde a integral trapezoidal da corrente ao longo do tempo.",
            normal,
        )
    )

    doc.build(story)
    return out.getvalue()
